"""Receipt service — build and dispatch order receipts after checkout.

Dispatch mode is driven by config.RECEIPT_MODE:
  "zebra" — ZPL receipt sent to Zebra GX420D via USB
  "pdf"   — PDF saved to config.RECEIPT_OUTPUT_DIR then opened (plain-text
             .txt fallback when reportlab is not installed)
  "none"  — silent; no receipt produced
"""
import os
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime

import config

# ── Unicode font registration (Cyrillic support) ──────────────────────────────
_FONT_REGULAR = "Helvetica"
_FONT_BOLD    = "Helvetica-Bold"

def _register_fonts() -> None:
    """Register a Unicode-capable TTF font for PDF output. No-op if already done."""
    global _FONT_REGULAR, _FONT_BOLD
    if _FONT_REGULAR != "Helvetica":
        return
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.fonts import addMapping

        candidates = [
            ("C:/Windows/Fonts/arial.ttf",   "C:/Windows/Fonts/arialbd.ttf"),
            ("C:/Windows/Fonts/calibri.ttf", "C:/Windows/Fonts/calibrib.ttf"),
            ("/Library/Fonts/Arial.ttf",     "/Library/Fonts/Arial Bold.ttf"),
        ]
        for reg, bold in candidates:
            if os.path.exists(reg) and os.path.exists(bold):
                pdfmetrics.registerFont(TTFont("AppFont",      reg))
                pdfmetrics.registerFont(TTFont("AppFont-Bold", bold))
                pdfmetrics.registerFontFamily(
                    "AppFont",
                    normal="AppFont",
                    bold="AppFont-Bold",
                    italic="AppFont",
                    boldItalic="AppFont-Bold",
                )
                _FONT_REGULAR = "AppFont"
                _FONT_BOLD    = "AppFont-Bold"
                break
    except Exception:
        pass


def _apply_font_to_styles(styles) -> None:
    """Override default Helvetica in reportlab stylesheet with the registered font."""
    for s in styles.byName.values():
        if hasattr(s, "fontName"):
            if "Bold" in (s.fontName or ""):
                s.fontName = _FONT_BOLD
            else:
                s.fontName = _FONT_REGULAR


# ── Physical receipt layout ───────────────────────────────────────────────────
# Measurements are defined once, in inches, then converted to printer dots
# (ZPL, via config.ZEBRA_DPI) or PDF points (72/in) below, so a receipt looks
# the same regardless of which renderer produced it or the printer's DPI.
_MARGIN_IN     = config.RECEIPT_WIDTH_IN * 0.10   # 10% of receipt width, left/right each
_TITLE_SIZE_IN = 0.177   # store name
_ADDR_SIZE_IN  = 0.108   # address / secondary lines
_BODY_SIZE_IN  = 0.128   # item lines, totals
_TOTAL_SIZE_IN = 0.148   # TOTAL emphasis line
_LINE_PITCH_IN = 0.167   # line-to-line vertical pitch
_SEP_PAD_IN    = 0.030   # padding around a separator rule
_TOTAL_GAP_IN  = _LINE_PITCH_IN       # extra breathing room before the TOTAL block (one blank line)
_TOP_GAP_IN    = 1.0                  # blank space at the very top of the receipt, in inches
_BOTTOM_GAP_IN = 0.5                  # blank space at the very bottom of the receipt, in inches

# Line pitch as a multiple of the line's own font size (derived from body
# text, where 0.167in pitch / 0.128in size ≈ 1.3x). Applying this ratio to
# every size — not just body — keeps a line's own font (ascent+descent)
# inside its pitch, so a taller line (e.g. the title) never crowds the
# line below it.
_LINE_PITCH_RATIO = _LINE_PITCH_IN / _BODY_SIZE_IN

_ZPL_WIDTH_DOTS = round(config.RECEIPT_WIDTH_IN * config.ZEBRA_DPI)
_MARGIN_X       = round(_MARGIN_IN     * config.ZEBRA_DPI)
_FONT_H         = round(_BODY_SIZE_IN  * config.ZEBRA_DPI)   # font height in dots (body text)
_LINE_H         = round(_LINE_PITCH_IN * config.ZEBRA_DPI)   # line pitch in dots
_SEP_Y_PAD      = round(_SEP_PAD_IN    * config.ZEBRA_DPI)
_TITLE_H        = round(_TITLE_SIZE_IN * config.ZEBRA_DPI)
_ADDR_H         = round(_ADDR_SIZE_IN  * config.ZEBRA_DPI)
_TOTAL_H        = round(_TOTAL_SIZE_IN * config.ZEBRA_DPI)
_TOTAL_GAP      = round(_TOTAL_GAP_IN  * config.ZEBRA_DPI)
_TOP_GAP        = round(_TOP_GAP_IN    * config.ZEBRA_DPI)   # top gap in dots (ZPL)
_TOP_GAP_PT     = _TOP_GAP_IN * 72                            # top gap in points (PDF)
_BOTTOM_GAP     = round(_BOTTOM_GAP_IN * config.ZEBRA_DPI)    # bottom gap in dots (ZPL)
_BOTTOM_GAP_PT  = _BOTTOM_GAP_IN * 72                         # bottom gap in points (PDF)

# ── Helpers ───────────────────────────────────────────────────────────────────

def _fmt_dt(iso: str) -> str:
    """ISO datetime → US display format ('06/21/2026 02:30 PM')."""
    try:
        return datetime.strptime(iso[:16], "%Y-%m-%d %H:%M").strftime("%m/%d/%Y %I:%M %p")
    except (ValueError, TypeError):
        return iso

# Name-wrap width in characters — sized to the printable width at a rough
# 0.6x-em average character width, shared by ZPL and the plain-text receipt.
_TXT_WIDTH = max(20, int((config.RECEIPT_WIDTH_IN - 2 * _MARGIN_IN) / (0.6 * _BODY_SIZE_IN)))


# ── Shared receipt content ────────────────────────────────────────────────────
# One instruction list describes what a receipt says; the PDF, ZPL-image, and
# ZPL-fields renderers each just interpret it with their own drawing
# primitives. A content change (new line, different conditional) only needs
# to happen here — it then applies to every renderer automatically. Item-name
# wrapping is renderer-specific (different font-metric engines: PIL vs
# reportlab vs character-count), so ItemLine keeps the raw name/qty/price
# and each renderer wraps it with its own method.

@dataclass
class Center:
    text: str
    size: str = "body"   # "title" | "addr" | "body" | "total"
    bold: bool = False


@dataclass
class Pair:
    """Left-aligned label + right-justified value on the same line."""
    left: str
    right: str
    size: str = "body"
    bold: bool = False


@dataclass
class ItemLine:
    """A line item — name (wrapped per-renderer) plus a qty/price Pair row."""
    name: str
    qty: int
    price: float


@dataclass
class Rule:
    thick: bool = False


@dataclass
class Gap:
    """Extra vertical breathing room (one _TOTAL_GAP)."""


def _receipt_content(order: dict, items: list[dict]) -> list:
    """Build the ordered content of a receipt, shared by every renderer."""
    is_gift  = order.get("payment_method") == "gift"
    sym      = config.CURRENCY_SYMBOL
    tax_rate = 0.0 if is_gift else config.TAX_RATE
    disc_pct = order.get("discount_pct") or 0.0
    subtotal = sum(i["quantity"] * i["unit_price"] for i in items)
    disc_amt = round(subtotal * disc_pct / 100, 2)
    taxable  = subtotal - disc_amt
    tax      = round(taxable * tax_rate, 2)
    total    = taxable + tax
    dt       = _fmt_dt(order.get("created_at", datetime.now().strftime("%Y-%m-%d %H:%M")))

    content: list = [Center(config.STORE_NAME, size="title", bold=True)]
    for addr_line in config.STORE_ADDRESS:
        content.append(Center(addr_line, size="addr"))

    if is_gift:
        content.append(Center("** GIFT RECEIPT **", size="total", bold=True))
        if order.get("customer_name"):
            content.append(Center(f"For: {order['customer_name']}"))

    content.append(Center(dt))
    content.append(Center(f"Order #{order['id']}"))
    if order.get("processed_by"):
        content.append(Center(f"Served by: {order['processed_by']}"))
    if not is_gift:
        content.append(Center(f"Payment: {(order.get('payment_method') or 'cash').upper()}"))

    content.append(Gap())
    content.append(Rule(thick=True))
    content.append(Gap())

    for item in items:
        content.append(ItemLine(item["product_name"], item["quantity"],
                                 item["quantity"] * item["unit_price"]))

    content.append(Rule())

    content.append(Pair("  Subtotal", f"{sym}{subtotal:.2f}"))
    if disc_pct:
        content.append(Pair(f"  Discount ({disc_pct:.4g}%)", f"-{sym}{disc_amt:.2f}"))
    if not is_gift:
        content.append(Pair("  Sales Tax", f"{sym}{tax:.2f}"))

    content.append(Gap())
    content.append(Rule(thick=True))
    content.append(Gap())
    content.append(Pair("  TOTAL", f"{sym}{total:.2f}", size="total", bold=True))
    content.append(Rule())
    content.append(Gap())
    content.append(Center("Thank you!"))

    return content


def _fit_name(name: str, max_chars: int) -> list[str]:
    """Fit a product name into at most 2 lines; truncate remainder with '...'."""
    upper = name.upper().strip()
    if not upper:
        return [""]
    if len(upper) <= max_chars:
        return [upper]
    words = upper.split()
    line1: list[str] = []
    for word in words:
        if len(" ".join(line1 + [word])) <= max_chars:
            line1.append(word)
        else:
            break
    l1 = " ".join(line1)
    l2 = " ".join(words[len(line1):])
    if not l1:
        return [upper[:max_chars - 3] + "..."]
    if not l2:
        return [l1]
    if len(l2) <= max_chars:
        return [l1, l2]
    return [l1, l2[:max_chars - 3] + "..."]


def _fit_name_px(name: str, font: str, size: float, max_w: float) -> list[str]:
    """Pixel-accurate 2-line fit for proportional fonts (reportlab)."""
    from reportlab.pdfbase.pdfmetrics import stringWidth as sw
    upper = name.upper().strip()
    if not upper:
        return [""]
    if sw(upper, font, size) <= max_w:
        return [upper]
    words = upper.split()
    line1: list[str] = []
    for word in words:
        if sw(" ".join(line1 + [word]), font, size) <= max_w:
            line1.append(word)
        else:
            break
    l1 = " ".join(line1)
    l2 = " ".join(words[len(line1):])
    if not l1:
        for i in range(len(upper) - 1, 0, -1):
            if sw(upper[:i] + "...", font, size) <= max_w:
                return [upper[:i] + "..."]
        return ["..."]
    if not l2:
        return [l1]
    if sw(l2, font, size) <= max_w:
        return [l1, l2]
    for i in range(len(l2) - 1, 0, -1):
        if sw(l2[:i] + "...", font, size) <= max_w:
            return [l1, l2[:i] + "..."]
    return [l1, "..."]


# ── Public API ────────────────────────────────────────────────────────────────

def print_receipt(order: dict, items: list[dict]) -> None:
    """
    Dispatch a receipt according to config.RECEIPT_MODE.
    Raises OSError / IOError on failure so the caller can show a warning.
    """
    mode = config.RECEIPT_MODE

    if mode == "zebra":
        from services.zebra_service import print_label_usb
        print_label_usb(build_receipt_zpl(order, items))

    elif mode == "pdf":
        os.makedirs(config.RECEIPT_OUTPUT_DIR, exist_ok=True)
        requested = os.path.join(
            config.RECEIPT_OUTPUT_DIR,
            f"receipt_{order['id']}.pdf",
        )
        actual_path = build_receipt_pdf(order, items, requested)
        _open_file(actual_path)

    # "none" → do nothing


# ── ZPL receipt ───────────────────────────────────────────────────────────────

def build_receipt_zpl(order: dict, items: list[dict]) -> str:
    """
    Build a ZPL receipt for the Zebra printer on 4" continuous paper.

    Renders the receipt as a bitmap (Pillow) and embeds it via ^GF —
    this uses real font metrics for alignment instead of ^A0N's
    proportional-font quirks. Falls back to plain ZPL field commands
    if Pillow isn't installed.
    """
    try:
        return _build_receipt_zpl_image(order, items)
    except ImportError:
        return _build_receipt_zpl_fields(order, items)


def _build_receipt_zpl_image(order: dict, items: list[dict]) -> str:
    """Render the receipt as a bitmap and embed it via ZPL's ^GF. Raises ImportError if Pillow is missing."""
    from PIL import Image, ImageDraw, ImageFont
    from services.zebra_service import find_ttf_fonts, image_to_zpl_gf

    content = _receipt_content(order, items)
    sym     = config.CURRENCY_SYMBOL

    reg_path, bold_path = find_ttf_fonts()

    def make_font(size: int, bold: bool = False):
        path = (bold_path or reg_path) if bold else reg_path
        return ImageFont.truetype(path, size) if path else ImageFont.load_default()

    size_px    = {"title": _TITLE_H, "addr": _ADDR_H, "body": _FONT_H, "total": _TOTAL_H}
    pitch_px   = {k: round(v * _LINE_PITCH_RATIO) for k, v in size_px.items()}
    font_cache: dict[tuple[str, bool], object] = {}

    def get_font(size_key: str, bold: bool):
        key = (size_key, bold)
        if key not in font_cache:
            font_cache[key] = make_font(size_px[size_key], bold=bold)
        return font_cache[key]

    width      = _ZPL_WIDTH_DOTS
    max_name_w = width - _MARGIN_X * 2

    # Measuring surface — draw.textlength only reads font metrics, so its
    # backing image size is irrelevant; used for both passes below.
    measure_draw = ImageDraw.Draw(Image.new("L", (1, 1), 255))

    def fit_name(name: str, fnt, max_w: int) -> list[str]:
        """Pixel-accurate 2-line fit, mirroring _fit_name_px but for a PIL font."""
        upper = name.upper().strip()
        if not upper:
            return [""]
        if measure_draw.textlength(upper, font=fnt) <= max_w:
            return [upper]
        words = upper.split()
        line1: list[str] = []
        for word in words:
            if measure_draw.textlength(" ".join(line1 + [word]), font=fnt) <= max_w:
                line1.append(word)
            else:
                break
        l1 = " ".join(line1)
        l2 = " ".join(words[len(line1):])
        if not l1:
            for i in range(len(upper) - 1, 0, -1):
                if measure_draw.textlength(upper[:i] + "...", font=fnt) <= max_w:
                    return [upper[:i] + "..."]
            return ["..."]
        if not l2:
            return [l1]
        if measure_draw.textlength(l2, font=fnt) <= max_w:
            return [l1, l2]
        for i in range(len(l2) - 1, 0, -1):
            if measure_draw.textlength(l2[:i] + "...", font=fnt) <= max_w:
                return [l1, l2[:i] + "..."]
        return [l1, "..."]

    # Pass 1: measure exact content height (item-name wrap included) instead
    # of guessing via a fixed canvas — a too-small guess let PIL's crop pad
    # the overflow with solid black across the full width (a big order's
    # receipt would print a black bar from wherever the guess ran out).
    y = 20 + _TOP_GAP
    for instr in content:
        if isinstance(instr, Center) or isinstance(instr, Pair):
            y += pitch_px[instr.size]
        elif isinstance(instr, Rule):
            thick = 2 if instr.thick else 1
            y = y + thick + _SEP_Y_PAD * 2
        elif isinstance(instr, Gap):
            y += _TOTAL_GAP
        elif isinstance(instr, ItemLine):
            fnt = get_font("body", False)
            y += _LINE_H * (len(fit_name(instr.name, fnt, max_name_w)) + 1)
    label_height = y + 20 + _BOTTOM_GAP   # bottom margin + bottom gap

    img  = Image.new("L", (width, label_height), 255)
    draw = ImageDraw.Draw(img)

    # Pass 2: draw.
    y = 20 + _TOP_GAP
    for instr in content:
        if isinstance(instr, Center):
            fnt = get_font(instr.size, instr.bold)
            w = draw.textlength(instr.text, font=fnt)
            draw.text(((width - w) / 2, y), instr.text, font=fnt, fill=0)
            y += pitch_px[instr.size]
        elif isinstance(instr, Pair):
            fnt = get_font(instr.size, instr.bold)
            draw.text((_MARGIN_X, y), instr.left, font=fnt, fill=0)
            rw = draw.textlength(instr.right, font=fnt)
            draw.text((width - _MARGIN_X - rw, y), instr.right, font=fnt, fill=0)
            y += pitch_px[instr.size]
        elif isinstance(instr, Rule):
            thick = 2 if instr.thick else 1
            top = y + _SEP_Y_PAD
            draw.rectangle([_MARGIN_X, top, width - _MARGIN_X, top + thick - 1], fill=0)
            y = y + thick + _SEP_Y_PAD * 2
        elif isinstance(instr, Gap):
            y += _TOTAL_GAP
        elif isinstance(instr, ItemLine):
            fnt = get_font("body", False)
            for nl in fit_name(instr.name, fnt, max_name_w):
                draw.text((_MARGIN_X, y), nl, font=fnt, fill=0)
                y += _LINE_H
            right = f"{sym}{instr.price:.2f}"
            draw.text((_MARGIN_X, y), f"  x{instr.qty}", font=fnt, fill=0)
            rw = draw.textlength(right, font=fnt)
            draw.text((width - _MARGIN_X - rw, y), right, font=fnt, fill=0)
            y += _LINE_H

    gf_cmd, gw, gh = image_to_zpl_gf(img)

    header = (
        "^XA\n"
        "^MNC\n"
        f"^PW{width}\n"
        f"^LL{gh}\n"
        "^CI28\n"
    )
    return header + f"^FO0,0{gf_cmd}^FS\n" + "^XZ\n"


def _build_receipt_zpl_fields(order: dict, items: list[dict]) -> str:
    """Fallback ZPL builder using plain field commands (no Pillow required)."""
    content = _receipt_content(order, items)
    sym     = config.CURRENCY_SYMBOL
    size_h  = {"title": _TITLE_H, "addr": _ADDR_H, "body": _FONT_H, "total": _TOTAL_H}
    block_w = _ZPL_WIDTH_DOTS - _MARGIN_X * 2

    lines: list[str] = []
    y = 20 + _TOP_GAP

    def emit_center(text: str, font_h: int) -> None:
        nonlocal y
        lines.append(f"^FO{_MARGIN_X},{y}^A0N,{font_h},{font_h}^FB{block_w},1,0,C^FD{text}^FS")
        y += round(font_h * _LINE_PITCH_RATIO)

    def emit_left(text: str, font_h: int) -> None:
        nonlocal y
        lines.append(f"^FO{_MARGIN_X},{y}^A0N,{font_h},{font_h}^FD{text}^FS")
        y += round(font_h * _LINE_PITCH_RATIO)

    def emit_pair(left: str, right: str, font_h: int) -> None:
        """^A0N is proportional, not monospace — ^FB's right-justify (not
        manual space-padding) is what actually lands the value at the
        right margin."""
        nonlocal y
        lines.append(f"^FO{_MARGIN_X},{y}^A0N,{font_h},{font_h}^FD{left}^FS")
        lines.append(f"^FO{_MARGIN_X},{y}^A0N,{font_h},{font_h}^FB{block_w},1,0,R^FD{right}^FS")
        y += round(font_h * _LINE_PITCH_RATIO)

    def emit_rule(thick: bool) -> None:
        nonlocal y
        t = 2 if thick else 1
        lines.append(f"^FO{_MARGIN_X},{y + _SEP_Y_PAD}^GB{block_w},{t},{t}^FS")
        y += t + _SEP_Y_PAD * 2

    for instr in content:
        if isinstance(instr, Center):
            emit_center(instr.text, size_h[instr.size])
        elif isinstance(instr, Pair):
            emit_pair(instr.left, instr.right, size_h[instr.size])
        elif isinstance(instr, Rule):
            emit_rule(instr.thick)
        elif isinstance(instr, Gap):
            y += _TOTAL_GAP
        elif isinstance(instr, ItemLine):
            for nl in _fit_name(instr.name, _TXT_WIDTH):
                emit_left(nl, _FONT_H)
            emit_pair(f"  x{instr.qty}", f"{sym}{instr.price:.2f}", _FONT_H)

    label_height = y + 20 + _BOTTOM_GAP   # bottom margin + bottom gap
    header = (
        "^XA\n"
        "^MNC\n"                    # continuous media (no gaps) for thermal tape
        f"^PW{_ZPL_WIDTH_DOTS}\n"
        f"^LL{label_height}\n"
        "^CI28\n"           # UTF-8
    )
    return header + "\n".join(lines) + "\n^XZ\n"


# ── PDF / txt receipt ─────────────────────────────────────────────────────────

def build_receipt_pdf(order: dict, items: list[dict], path: str) -> str:
    """
    Write a receipt and return the actual file path created.
    Uses reportlab for a proper PDF; falls back to a .txt file.
    """
    try:
        _build_pdf_reportlab(order, items, path)
        return path
    except ImportError:
        txt_path = os.path.splitext(path)[0] + ".txt"
        _build_txt(order, items, txt_path)
        return txt_path


def _build_txt(order: dict, items: list[dict], path: str) -> None:
    is_gift  = order.get("payment_method") == "gift"
    sym      = config.CURRENCY_SYMBOL
    tax_rate = 0.0 if is_gift else config.TAX_RATE
    disc_pct = order.get("discount_pct") or 0.0
    subtotal = sum(i["quantity"] * i["unit_price"] for i in items)
    disc_amt = round(subtotal * disc_pct / 100, 2)
    taxable  = subtotal - disc_amt
    tax      = round(taxable * tax_rate, 2)
    total    = taxable + tax
    w        = _TXT_WIDTH
    dt       = _fmt_dt(order.get("created_at", datetime.now().strftime("%Y-%m-%d %H:%M")))

    sep_thick = "=" * w
    sep_thin  = "-" * w

    receipt_lines = [
        sep_thick,
        config.STORE_NAME.center(w),
    ]
    for addr_line in config.STORE_ADDRESS:
        receipt_lines.append(addr_line.center(w))
    if is_gift:
        receipt_lines.append("** GIFT RECEIPT **".center(w))
        if order.get("customer_name"):
            receipt_lines.append(f"For: {order['customer_name']}".center(w))
    receipt_lines.append(dt.center(w))
    receipt_lines.append(f"Order #{order['id']}".center(w))
    if order.get("processed_by"):
        receipt_lines.append(f"Served by: {order['processed_by']}".center(w))
    if not is_gift:
        receipt_lines.append(f"Payment: {(order.get('payment_method') or 'cash').upper()}".center(w))
    receipt_lines.append(sep_thick)

    for item in items:
        name_lines = _fit_name(item["product_name"], w)
        qty   = item["quantity"]
        price = item["quantity"] * item["unit_price"]
        right = f"{sym}{price:.2f}"
        receipt_lines.extend(name_lines)
        gap = w - 2 - len(f"x{qty}") - len(right)
        receipt_lines.append(f"  x{qty}" + " " * max(gap, 1) + right)

    disc_lines = (
        [_total_line(f"Discount ({disc_pct:.4g}%)", f"-{sym}{disc_amt:.2f}", w)]
        if disc_pct else []
    )
    tax_lines = [] if is_gift else [_total_line("Sales Tax", f"{sym}{tax:.2f}", w)]
    receipt_lines += [
        sep_thin,
        _total_line("Subtotal", f"{sym}{subtotal:.2f}", w),
        *disc_lines,
        *tax_lines,
        sep_thick,
        "",
        _total_line("TOTAL", f"{sym}{total:.2f}", w),
        sep_thick,
        "",
        "Thank you!".center(w),
        "",
    ]

    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(receipt_lines))


def _total_line(label: str, value: str, width: int) -> str:
    gap = width - len(label) - len(value) - 2
    return f"  {label}" + " " * max(gap, 1) + value


def _build_pdf_reportlab(order: dict, items: list[dict], path: str) -> None:
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas as rl_canvas
    _register_fonts()

    content = _receipt_content(order, items)
    sym     = config.CURRENCY_SYMBOL

    # Page matches the physical receipt paper width (same constants as ZPL).
    title_sz = round(_TITLE_SIZE_IN * 72, 1)
    addr_sz  = round(_ADDR_SIZE_IN  * 72, 1)
    body_sz  = round(_BODY_SIZE_IN  * 72, 1)
    total_sz = round(_TOTAL_SIZE_IN * 72, 1)
    size_map = {"title": title_sz, "addr": addr_sz, "body": body_sz, "total": total_sz}

    page_w     = config.RECEIPT_WIDTH_IN * inch
    line_h     = _LINE_PITCH_IN * 72
    margin     = _MARGIN_IN * inch
    max_name_w = page_w - margin * 2

    # Pass 1: measure exact content height (item-name wrap included) instead
    # of guessing via a hand-tuned line-count formula — that's what let the
    # "Thank you!" line silently fall off the bottom of the page earlier.
    content_h = 0.0
    for instr in content:
        if isinstance(instr, ItemLine):
            name_lines = _fit_name_px(instr.name, _FONT_REGULAR, body_sz, max_name_w)
            content_h += line_h * (len(name_lines) + 1)
        elif isinstance(instr, Rule):
            content_h += 4
        elif isinstance(instr, Gap):
            content_h += _TOTAL_GAP_IN * 72
        else:
            content_h += size_map[instr.size] * _LINE_PITCH_RATIO
    page_h = content_h + 40 + _TOP_GAP_PT + _BOTTOM_GAP_PT   # 20pt top margin + 20pt bottom margin + top/bottom gaps

    c = rl_canvas.Canvas(path, pagesize=(page_w, page_h))
    y = page_h - 20 - _TOP_GAP_PT

    def draw_center(text: str, size: float, bold: bool = False):
        nonlocal y
        c.setFont(_FONT_BOLD if bold else _FONT_REGULAR, size)
        c.drawCentredString(page_w / 2, y, text)
        y -= size * _LINE_PITCH_RATIO

    def draw_pair(left: str, right: str, size: float, bold: bool = False):
        nonlocal y
        c.setFont(_FONT_BOLD if bold else _FONT_REGULAR, size)
        c.drawString(margin, y, left)
        c.drawRightString(page_w - margin, y, right)
        y -= size * _LINE_PITCH_RATIO

    def draw_rule(thick: bool):
        nonlocal y
        c.setLineWidth(1.5 if thick else 0.5)
        c.line(margin, y + line_h * 0.4, page_w - margin, y + line_h * 0.4)
        y -= 4

    # Pass 2: draw.
    for instr in content:
        if isinstance(instr, Center):
            draw_center(instr.text, size_map[instr.size], instr.bold)
        elif isinstance(instr, Pair):
            draw_pair(instr.left, instr.right, size_map[instr.size], instr.bold)
        elif isinstance(instr, Rule):
            draw_rule(instr.thick)
        elif isinstance(instr, Gap):
            y -= _TOTAL_GAP_IN * 72
        elif isinstance(instr, ItemLine):
            c.setFont(_FONT_REGULAR, body_sz)
            for nl in _fit_name_px(instr.name, _FONT_REGULAR, body_sz, max_name_w):
                c.drawString(margin, y, nl)
                y -= line_h
            draw_pair(f"  x{instr.qty}", f"{sym}{instr.price:.2f}", body_sz)

    c.save()


# ── Detailed sales report ────────────────────────────────────────────────────

def _build_detailed_pdf_reportlab(orders, date_from, date_to, path, now):
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet
    _register_fonts()
    styles = getSampleStyleSheet()
    _apply_font_to_styles(styles)

    sym    = config.CURRENCY_SYMBOL
    period = f"{date_from or 'All'} – {date_to or 'All'}"

    doc = SimpleDocTemplate(path, pagesize=LETTER,
                            leftMargin=0.6*inch, rightMargin=0.6*inch,
                            topMargin=0.5*inch,  bottomMargin=0.5*inch)

    title_style  = styles["Title"]
    normal       = styles["Normal"]
    normal.fontSize = 9

    bold_style = styles["Normal"].__class__(
        "Bold9", parent=styles["Normal"],
        fontName=_FONT_BOLD, fontSize=9,
    )

    story = []
    story.append(Paragraph("DETAILED SALES REPORT", title_style))
    story.append(Paragraph(config.STORE_NAME, normal))
    story.append(Paragraph(f"Period: {period}", normal))
    story.append(Paragraph(_fmt_dt(now.strftime("%Y-%m-%d %H:%M")), normal))
    story.append(Spacer(1, 0.15*inch))

    from models import order as order_model

    grand_sub = grand_disc = grand_tax = grand_total = 0.0

    for o in orders:
        items   = order_model.get_items(o["id"])
        sub     = sum(i["quantity"] * i["unit_price"] for i in items)
        disc    = o.get("discount_pct") or 0
        disc_amt = round(sub * disc / 100, 2)
        tax     = round((sub - disc_amt) * config.TAX_RATE, 2)
        total   = o["total"]
        method  = (o.get("payment_method") or "cash").upper()
        by      = o.get("processed_by") or "—"

        grand_sub   += sub
        grand_disc  += disc_amt
        grand_tax   += tax
        grand_total += total

        # Order header row
        hdr_data = [[
            Paragraph(f"<b>Order #{o['id']}</b>", normal),
            Paragraph(f"<b>{_fmt_dt(o['created_at'])}</b>", normal),
            Paragraph(f"<b>{method}</b>", normal),
            Paragraph(f"<b>By: {by}</b>", normal),
            Paragraph(f"<b>Total: {sym}{total:.2f}</b>", normal),
        ]]
        hdr_tbl = Table(hdr_data, colWidths=[1.1*inch, 1.8*inch, 0.7*inch, 1.5*inch, 1.3*inch])
        hdr_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#2980B9")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), _FONT_BOLD),
            ("FONTSIZE",    (0, 0), (-1, 0), 9),
            ("TOPPADDING",  (0, 0), (-1, 0), 4),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
            ("VALIGN",      (0, 0), (-1, 0), "MIDDLE"),
        ]))
        story.append(hdr_tbl)

        # Items sub-table
        item_rows = [["Product", "Qty", "Unit Price", "Line Total"]]
        for item in items:
            line = item["quantity"] * item["unit_price"]
            item_rows.append([
                Paragraph(item["product_name"], normal),
                str(item["quantity"]),
                f"{sym}{item['unit_price']:.2f}",
                f"{sym}{line:.2f}",
            ])

        # Subtotals row
        disc_str = f"  disc {disc:.4g}%  −{sym}{disc_amt:.2f}" if disc else ""
        item_rows.append([
            Paragraph(f"<b>Subtotal{disc_str}</b>", normal),
            "", f"Tax {sym}{tax:.2f}",
            Paragraph(f"<b>{sym}{total:.2f}</b>", normal),
        ])

        item_tbl = Table(item_rows,
                         colWidths=[3.4*inch, 0.5*inch, 1.2*inch, 1.3*inch],
                         repeatRows=0)
        item_tbl.setStyle(TableStyle([
            ("FONTNAME",       (0, 0),  (-1, 0),  _FONT_BOLD),
            ("FONTNAME",       (0, 1),  (-1, -1), _FONT_REGULAR),
            ("FONTSIZE",       (0, 0),  (-1, -1), 8),
            ("BACKGROUND",     (0, 0),  (-1, 0),  colors.HexColor("#ECF0F1")),
            ("ROWBACKGROUNDS", (0, 1),  (-1, -2),
             [colors.white, colors.HexColor("#EAF4FB")]),
            ("BACKGROUND",     (0, -1), (-1, -1), colors.HexColor("#ECF0F1")),
            ("FONTNAME",       (0, -1), (-1, -1), _FONT_BOLD),
            ("ALIGN",          (1, 0),  (-1, -1), "RIGHT"),
            ("GRID",           (0, 0),  (-1, -1), 0.25, colors.HexColor("#BDC3C7")),
            ("LEFTPADDING",    (0, 0),  (0, -1),  6),
            ("TOPPADDING",     (0, 0),  (-1, -1), 3),
            ("BOTTOMPADDING",  (0, 0),  (-1, -1), 3),
            ("VALIGN",         (0, 0),  (-1, -1), "MIDDLE"),
        ]))
        story.append(item_tbl)
        story.append(Spacer(1, 0.12*inch))

    # Grand totals
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2980B9")))
    story.append(Spacer(1, 0.06*inch))
    story.append(Paragraph(
        f"Orders: <b>{len(orders)}</b> &nbsp;&nbsp; "
        f"Subtotal: <b>{sym}{grand_sub:.2f}</b> &nbsp;&nbsp; "
        f"Discounts: <b>−{sym}{grand_disc:.2f}</b> &nbsp;&nbsp; "
        f"Tax: <b>{sym}{grand_tax:.2f}</b> &nbsp;&nbsp; "
        f"Total Revenue: <b>{sym}{grand_total:.2f}</b>",
        normal,
    ))
    doc.build(story)


def build_detailed_report_pdf(orders: list[dict],
                               date_from: str | None,
                               date_to:   str | None) -> str:
    """Generate a detailed per-order report PDF, open it, return the path."""
    os.makedirs(config.RECEIPT_OUTPUT_DIR, exist_ok=True)
    now  = datetime.now()
    stem = f"detailed_report_{now.strftime('%Y%m%d_%H%M%S')}"
    pdf_path = os.path.join(config.RECEIPT_OUTPUT_DIR, stem + ".pdf")
    try:
        _build_detailed_pdf_reportlab(orders, date_from, date_to, pdf_path, now)
        _open_file(pdf_path)
        return pdf_path
    except ImportError:
        raise RuntimeError("reportlab is required for PDF reports.")


def build_detailed_report_xlsx(orders: list[dict],
                                date_from: str | None,
                                date_to:   str | None) -> str:
    """Generate a detailed per-order report XLSX, open it, and return the path."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from models import order as order_model

    os.makedirs(config.RECEIPT_OUTPUT_DIR, exist_ok=True)
    now  = datetime.now()
    stem = f"detailed_report_{now.strftime('%Y%m%d_%H%M%S')}"
    path = os.path.join(config.RECEIPT_OUTPUT_DIR, stem + ".xlsx")
    sym    = config.CURRENCY_SYMBOL
    period = f"{date_from or 'All'} - {date_to or 'All'}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detailed Report"

    ws.append(["DETAILED SALES REPORT"])
    ws.append([config.STORE_NAME])
    ws.append([f"Period: {period}"])
    ws.append([_fmt_dt(now.strftime("%Y-%m-%d %H:%M"))])
    ws.append([])

    hdr_fill  = PatternFill("solid", fgColor="2980B9")
    item_fill = PatternFill("solid", fgColor="ECF0F1")
    hdr_font  = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    grand_sub = grand_disc = grand_tax = grand_total = 0.0

    for o in orders:
        items    = order_model.get_items(o["id"])
        sub      = sum(i["quantity"] * i["unit_price"] for i in items)
        disc     = o.get("discount_pct") or 0
        disc_amt = round(sub * disc / 100, 2)
        tax      = round((sub - disc_amt) * config.TAX_RATE, 2)
        total    = o["total"]
        method   = (o.get("payment_method") or "cash").upper()
        by       = o.get("processed_by") or "—"

        grand_sub   += sub
        grand_disc  += disc_amt
        grand_tax   += tax
        grand_total += total

        ws.append([f"Order #{o['id']}", _fmt_dt(o["created_at"]), method,
                   f"By: {by}", f"Total: {sym}{total:.2f}"])
        for cell in ws[ws.max_row]:
            cell.fill = hdr_fill
            cell.font = hdr_font

        ws.append(["Product", "Qty", "Unit Price", "Line Total"])
        for cell in ws[ws.max_row]:
            cell.fill = item_fill
            cell.font = bold_font

        for item in items:
            line = item["quantity"] * item["unit_price"]
            ws.append([item["product_name"], item["quantity"],
                       item["unit_price"], line])

        disc_str = f"  disc {disc:.4g}%  -{sym}{disc_amt:.2f}" if disc else ""
        ws.append([f"Subtotal{disc_str}", "", f"Tax {sym}{tax:.2f}", total])
        for cell in ws[ws.max_row]:
            cell.font = bold_font
        ws.append([])

    ws.append(["", "", "Orders:", len(orders)])
    ws.append(["", "", "Subtotal:", grand_sub])
    ws.append(["", "", "Discounts:", -grand_disc])
    ws.append(["", "", "Tax:", grand_tax])
    ws.append(["", "", "Total Revenue:", grand_total])
    for r in range(ws.max_row - 4, ws.max_row + 1):
        for cell in ws[r]:
            cell.font = bold_font

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(path)
    _open_file(path)
    return path


# ── Sales report ─────────────────────────────────────────────────────────────

def build_sales_report_pdf(orders: list[dict],
                            date_from: str | None,
                            date_to:   str | None) -> str:
    """
    Generate a sales report PDF for the given orders, open it, and return
    the file path (.pdf or .txt fallback).
    """
    os.makedirs(config.RECEIPT_OUTPUT_DIR, exist_ok=True)
    now  = datetime.now()
    stem = f"sales_report_{now.strftime('%Y%m%d_%H%M%S')}"
    pdf_path = os.path.join(config.RECEIPT_OUTPUT_DIR, stem + ".pdf")

    try:
        _build_sales_pdf_reportlab(orders, date_from, date_to, pdf_path, now)
        actual = pdf_path
    except ImportError:
        txt_path = os.path.join(config.RECEIPT_OUTPUT_DIR, stem + ".txt")
        _build_sales_txt(orders, date_from, date_to, txt_path, now)
        actual = txt_path

    _open_file(actual)
    return actual


def build_sales_report_xlsx(orders: list[dict],
                             date_from: str | None,
                             date_to:   str | None) -> str:
    """Generate a sales report XLSX for the given orders, open it, and return the path."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    os.makedirs(config.RECEIPT_OUTPUT_DIR, exist_ok=True)
    now  = datetime.now()
    stem = f"sales_report_{now.strftime('%Y%m%d_%H%M%S')}"
    path = os.path.join(config.RECEIPT_OUTPUT_DIR, stem + ".xlsx")
    period = f"{date_from or 'All'} - {date_to or 'All'}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    ws.append(["SALES REPORT"])
    ws.append([config.STORE_NAME])
    ws.append([f"Period: {period}"])
    ws.append([_fmt_dt(now.strftime("%Y-%m-%d %H:%M"))])
    ws.append([])

    ws.append(["ID", "Date / Time", "Processed By", "Payment", "Items",
               "Subtotal", "Discount", "Tax", "Total"])
    hdr_fill = PatternFill("solid", fgColor="2980B9")
    hdr_font = Font(color="FFFFFF", bold=True)
    for cell in ws[ws.max_row]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    tot_items = tot_sub = tot_disc = tot_tax = tot_total = 0.0
    for o in orders:
        cnt, sub, disc, tax, total = _sales_order_row(o)
        method = (o.get("payment_method") or "cash").upper()
        by     = o.get("processed_by") or "—"
        dt     = _fmt_dt(o.get("created_at") or "")
        ws.append([o["id"], dt, by, method, cnt, sub, disc, tax, total])
        tot_items += cnt; tot_sub += sub; tot_disc += disc
        tot_tax   += tax; tot_total += total

    ws.append([f"{len(orders)} orders", "", "", "", tot_items,
               tot_sub, tot_disc, tot_tax, tot_total])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(path)
    _open_file(path)
    return path


def _sales_order_row(order: dict) -> tuple:
    """Return (items, subtotal, discount_amt, tax, total) for one order."""
    from models.order import get_items
    items    = get_items(order["id"])
    is_gift  = order.get("payment_method") == "gift"
    if is_gift:
        item_count = sum(i["quantity"] for i in items)
        return item_count, 0.0, 0.0, 0.0, 0.0
    subtotal    = sum(i["quantity"] * i["unit_price"] for i in items)
    disc_pct    = order.get("discount_pct") or 0.0
    disc_amt    = round(subtotal * disc_pct / 100, 2)
    taxable     = subtotal - disc_amt
    tax         = round(taxable * config.TAX_RATE, 2)
    total       = taxable + tax
    item_count  = sum(i["quantity"] for i in items)
    return item_count, subtotal, disc_amt, tax, total


def _build_sales_txt(orders, date_from, date_to, path, now):
    sym = config.CURRENCY_SYMBOL
    w   = 100
    period = f"{date_from or 'All'} – {date_to or 'All'}"

    lines = [
        "=" * w,
        "SALES REPORT".center(w),
        config.STORE_NAME.center(w),
        f"Period: {period}".center(w),
        _fmt_dt(now.strftime("%Y-%m-%d %H:%M")).center(w),
        "=" * w,
        f"{'ID':>5}  {'Date/Time':<18}  {'By':<14}  {'Pay':<5}  "
        f"{'Items':>5}  {'Subtotal':>9}  {'Disc':>8}  {'Tax':>8}  {'Total':>9}",
        "-" * w,
    ]

    tot_items = tot_sub = tot_disc = tot_tax = tot_total = 0.0
    for o in orders:
        cnt, sub, disc, tax, total = _sales_order_row(o)
        method = (o.get("payment_method") or "cash").upper()[:5]
        by     = (o.get("processed_by") or "")[:14]
        dt     = (o.get("created_at") or "")[:16]
        lines.append(
            f"{o['id']:>5}  {dt:<18}  {by:<14}  {method:<5}  "
            f"{cnt:>5}  {sym}{sub:>8.2f}  {sym}{disc:>7.2f}  "
            f"{sym}{tax:>7.2f}  {sym}{total:>8.2f}"
        )
        tot_items += cnt; tot_sub += sub; tot_disc += disc
        tot_tax   += tax; tot_total += total

    lines += [
        "=" * w,
        f"{'Orders: ' + str(len(orders)):>5}  {'':18}  {'':14}  {'':5}  "
        f"{tot_items:>5}  {sym}{tot_sub:>8.2f}  {sym}{tot_disc:>7.2f}  "
        f"{sym}{tot_tax:>7.2f}  {sym}{tot_total:>8.2f}",
        "=" * w, "",
    ]
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines))


def _build_sales_pdf_reportlab(orders, date_from, date_to, path, now):
    from reportlab.lib.pagesizes import LETTER, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    _register_fonts()
    from reportlab.lib.styles import getSampleStyleSheet

    styles = getSampleStyleSheet()
    _apply_font_to_styles(styles)

    sym    = config.CURRENCY_SYMBOL
    period = f"{date_from or 'All'} – {date_to or 'All'}"

    doc    = SimpleDocTemplate(path, pagesize=landscape(LETTER),
                               leftMargin=0.5*inch, rightMargin=0.5*inch,
                               topMargin=0.5*inch,  bottomMargin=0.5*inch)
    cs       = styles["Normal"]
    cs.fontSize = 9
    story  = []

    story.append(Paragraph("SALES REPORT", styles["Title"]))
    story.append(Paragraph(config.STORE_NAME, styles["Normal"]))
    story.append(Paragraph(f"Period: {period}", styles["Normal"]))
    story.append(Paragraph(_fmt_dt(now.strftime("%Y-%m-%d %H:%M")), styles["Normal"]))
    story.append(Spacer(1, 0.2*inch))

    header = ["ID", "Date / Time", "Processed By", "Payment", "Items",
              "Subtotal", "Discount", "Tax", "Total"]
    rows   = [header]

    tot_items = tot_sub = tot_disc = tot_tax = tot_total = 0.0
    for o in orders:
        cnt, sub, disc, tax, total = _sales_order_row(o)
        method  = (o.get("payment_method") or "cash").upper()
        by      = o.get("processed_by") or "—"
        dt      = _fmt_dt(o.get("created_at") or "")
        rows.append([
            str(o["id"]), dt, Paragraph(by, cs), method,
            str(int(cnt)),
            f"{sym}{sub:.2f}", f"{sym}{disc:.2f}",
            f"{sym}{tax:.2f}", f"{sym}{total:.2f}",
        ])
        tot_items += cnt; tot_sub += sub; tot_disc += disc
        tot_tax   += tax; tot_total += total

    rows.append([
        f"{len(orders)} orders", "", "", "", str(int(tot_items)),
        f"{sym}{tot_sub:.2f}", f"{sym}{tot_disc:.2f}",
        f"{sym}{tot_tax:.2f}", f"{sym}{tot_total:.2f}",
    ])

    col_widths = [0.45*inch, 1.5*inch, 1.4*inch, 0.8*inch, 0.5*inch,
                  0.85*inch, 0.85*inch, 0.75*inch, 0.85*inch]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0),  (-1, 0),   colors.HexColor("#2980B9")),
        ("TEXTCOLOR",      (0, 0),  (-1, 0),   colors.white),
        ("FONTNAME",       (0, 0),  (-1, 0),   _FONT_BOLD),
        ("FONTNAME",       (0, 1),  (-1, -1),  _FONT_REGULAR),
        ("FONTSIZE",       (0, 0),  (-1, -1),  9),
        ("ROWBACKGROUNDS", (0, 1),  (-1, -2),
         [colors.white, colors.HexColor("#EAF4FB")]),
        ("BACKGROUND",     (0, -1), (-1, -1),  colors.HexColor("#ECF0F1")),
        ("FONTNAME",       (0, -1), (-1, -1),  _FONT_BOLD),
        ("ALIGN",          (4, 1),  (-1, -1),  "RIGHT"),
        ("ALIGN",          (0, 0),  (0, -1),   "CENTER"),
        ("GRID",           (0, 0),  (-1, -1),  0.25, colors.HexColor("#BDC3C7")),
        ("VALIGN",         (0, 0),  (-1, -1),  "MIDDLE"),
        ("ROWHEIGHT",      (0, 0),  (-1, -1),  18),
    ]))
    story.append(tbl)

    story.append(Spacer(1, 0.15*inch))
    story.append(Paragraph(
        f"Orders: <b>{len(orders)}</b> &nbsp;&nbsp; "
        f"Items sold: <b>{int(tot_items)}</b> &nbsp;&nbsp; "
        f"Revenue: <b>{sym}{tot_total:.2f}</b> &nbsp;&nbsp; "
        f"Tax collected: <b>{sym}{tot_tax:.2f}</b>",
        styles["Normal"],
    ))
    doc.build(story)


# ── OS helpers ────────────────────────────────────────────────────────────────

def _open_file(path: str) -> None:
    """Open *path* with the OS default application."""
    if sys.platform == "win32":
        os.startfile(path)
    elif sys.platform == "darwin":
        subprocess.run(["open", path], check=False)
    else:
        subprocess.run(["xdg-open", path], check=False)


# ── Stock report ──────────────────────────────────────────────────────────────

def build_stock_report_pdf(products: list[dict]) -> str:
    """
    Generate a stock report, save to RECEIPT_OUTPUT_DIR, open it, and
    return the actual file path created (.pdf or .txt fallback).
    """
    os.makedirs(config.RECEIPT_OUTPUT_DIR, exist_ok=True)
    now  = datetime.now()
    stem = f"stock_report_{now.strftime('%Y%m%d_%H%M%S')}"

    pdf_path = os.path.join(config.RECEIPT_OUTPUT_DIR, stem + ".pdf")
    try:
        _build_stock_pdf_reportlab(products, pdf_path, now)
        actual = pdf_path
    except ImportError:
        txt_path = os.path.join(config.RECEIPT_OUTPUT_DIR, stem + ".txt")
        _build_stock_txt(products, txt_path, now)
        actual = txt_path

    _open_file(actual)
    return actual


def _build_stock_txt(products: list[dict], path: str, now: datetime) -> None:
    sym = config.CURRENCY_SYMBOL
    w   = 72

    total_units = sum(p["stock"] for p in products)
    total_value = sum(p["stock"] * p["price"] for p in products)

    lines = [
        "=" * w,
        "STOCK REPORT".center(w),
        config.STORE_NAME.center(w),
        _fmt_dt(now.strftime("%Y-%m-%d %H:%M")).center(w),
        "=" * w,
        f"{'ID':<6} {'Name':<28} {'Barcode':<12} {'Stock':>6} {'Price':>8} {'Value':>10}",
        "-" * w,
    ]

    for p in products:
        value      = p["stock"] * p["price"]
        stock_cell = str(p["stock"])
        lines.append(
            f"{p['id']:<6} {p['title'][:28]:<28} {p['barcode']:<12}"
            f" {stock_cell:>6} {sym}{p['price']:>7.2f} {sym}{value:>9.2f}"
        )

    lines += [
        "=" * w,
        f"  Products: {len(products)}   "
        f"Total units: {total_units}   "
        f"Stock value: {sym}{total_value:.2f}",
        "=" * w,
        "",
    ]

    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines))


def _build_stock_pdf_reportlab(products: list[dict], path: str,
                                now: datetime) -> None:
    from reportlab.lib.pagesizes import LETTER, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet
    _register_fonts()
    styles = getSampleStyleSheet()
    _apply_font_to_styles(styles)

    sym         = config.CURRENCY_SYMBOL
    total_units = sum(p["stock"] for p in products)
    total_value = sum(p["stock"] * p["price"] for p in products)

    doc    = SimpleDocTemplate(path, pagesize=landscape(LETTER),
                               leftMargin=0.5*inch, rightMargin=0.5*inch,
                               topMargin=0.5*inch,  bottomMargin=0.5*inch)
    story  = []

    # Header
    story.append(Paragraph("STOCK REPORT", styles["Title"]))
    story.append(Paragraph(config.STORE_NAME, styles["Normal"]))
    story.append(Paragraph(_fmt_dt(now.strftime("%Y-%m-%d %H:%M")),
                            styles["Normal"]))
    story.append(Spacer(1, 0.2*inch))

    cell_style = styles["Normal"]
    cell_style.fontSize = 9

    # Table
    header = ["ID", "Name", "Barcode", "Stock", "Price", "Value"]
    rows   = [header]
    for p in products:
        value      = p["stock"] * p["price"]
        stock_cell = str(p["stock"])
        rows.append([
            str(p["id"]),
            Paragraph(p["title"], cell_style),
            p["barcode"],
            stock_cell,
            f"{sym}{p['price']:.2f}",
            f"{sym}{value:.2f}",
        ])

    # Totals row
    rows.append([
        "", Paragraph("<b>TOTAL</b>", cell_style), "",
        str(total_units), "",
        f"{sym}{total_value:.2f}",
    ])

    col_widths = [0.5*inch, 3*inch, 1.2*inch, 0.7*inch, 0.9*inch, 0.9*inch]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#2980B9")),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",    (0, 0), (-1, 0),  _FONT_BOLD),
        ("FONTNAME",    (0, 1), (-1, -1), _FONT_REGULAR),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2),
         [colors.white, colors.HexColor("#EAF4FB")]),
        ("BACKGROUND",  (0, -1), (-1, -1), colors.HexColor("#ECF0F1")),
        ("FONTNAME",    (0, -1), (-1, -1), _FONT_BOLD),
        ("ALIGN",       (3, 1), (-1, -1), "RIGHT"),
        ("ALIGN",       (0, 0), (0, -1),  "CENTER"),
        ("GRID",        (0, 0), (-1, -1), 0.25, colors.HexColor("#BDC3C7")),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWHEIGHT",   (0, 0), (-1, -1), 18),
    ]))
    story.append(tbl)

    # Footer summary
    story.append(Spacer(1, 0.15*inch))
    story.append(Paragraph(
        f"Products: <b>{len(products)}</b> &nbsp;&nbsp; "
        f"Total units: <b>{total_units}</b> &nbsp;&nbsp; "
        f"Stock value: <b>{sym}{total_value:.2f}</b>",
        styles["Normal"],
    ))

    doc.build(story)

