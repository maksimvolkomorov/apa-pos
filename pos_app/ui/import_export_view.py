"""Import / Export page — xlsx only, column mapping via import_export_config.json."""
import json
import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox

import config
from models import product as product_model
from ui.theme import BG, BTN_BG, BTN_OK, BTN_DNG, FG_MUTED, styled_button


def _load_cfg() -> dict:
    path = os.path.join(config._BUNDLED, "import_export_config.json")
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


class ImportExportView(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG)
        self._build()

    # ── Layout ────────────────────────────────────────────────────────────────

    def _build(self):
        outer = tk.Frame(self, bg=BG)
        outer.pack(expand=True, anchor="n", fill="x", padx=60, pady=30)

        self._build_import_section(outer)
        tk.Frame(outer, bg="#CCCCCC", height=1).pack(fill="x", pady=24)
        self._build_export_section(outer)

        # Log area
        tk.Label(outer, text="Log", bg=BG,
                 font=("Helvetica", 10, "bold")).pack(anchor="w", pady=(16, 4))
        log_frame = tk.Frame(outer, bg=BG)
        log_frame.pack(fill="x")
        self._log = tk.Text(log_frame, height=10, state="disabled",
                            font=("Courier", 9), bg="#F7F7F7",
                            relief="solid", bd=1, wrap="word")
        sb = tk.Scrollbar(log_frame, command=self._log.yview)
        self._log.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._log.pack(side="left", fill="x", expand=True)

    def _build_import_section(self, parent):
        tk.Label(parent, text="Import from Excel (.xlsx)", bg=BG,
                 font=("Helvetica", 13, "bold")).pack(anchor="w")
        tk.Label(parent,
                 text="Column mapping is defined in import_export_config.json",
                 bg=BG, font=("Helvetica", 9), fg=FG_MUTED).pack(anchor="w", pady=(0, 10))

        row = tk.Frame(parent, bg=BG)
        row.pack(fill="x")

        self._import_path_var = tk.StringVar()
        tk.Entry(row, textvariable=self._import_path_var, width=55,
                 font=("Helvetica", 10), relief="solid", bd=1,
                 state="readonly").pack(side="left", padx=(0, 8))
        styled_button(row, "Browse…", self._browse_import).pack(side="left", padx=4)

        opt_row = tk.Frame(parent, bg=BG)
        opt_row.pack(anchor="w", pady=(10, 0))
        self._replace_var = tk.BooleanVar(value=True)
        tk.Checkbutton(opt_row, text="Replace all existing products",
                       variable=self._replace_var,
                       bg=BG, font=("Helvetica", 10),
                       activebackground=BG).pack(side="left")

        btn_row = tk.Frame(parent, bg=BG)
        btn_row.pack(anchor="w", pady=(12, 0))
        self._import_btn = styled_button(btn_row, "Import", self._run_import,
                                         bg=BTN_OK)
        self._import_btn.pack(side="left")
        self._import_btn.config(state="disabled")

        self._import_status = tk.Label(parent, text="", bg=BG,
                                        font=("Helvetica", 9), fg=FG_MUTED)
        self._import_status.pack(anchor="w", pady=(6, 0))

    def _build_export_section(self, parent):
        tk.Label(parent, text="Export to Excel (.xlsx)", bg=BG,
                 font=("Helvetica", 13, "bold")).pack(anchor="w")
        tk.Label(parent,
                 text="Exports all products from the database",
                 bg=BG, font=("Helvetica", 9), fg=FG_MUTED).pack(anchor="w", pady=(0, 10))

        btn_row = tk.Frame(parent, bg=BG)
        btn_row.pack(anchor="w")
        styled_button(btn_row, "Export…", self._run_export, bg=BTN_OK).pack(side="left")

        self._export_status = tk.Label(parent, text="", bg=BG,
                                        font=("Helvetica", 9), fg=FG_MUTED)
        self._export_status.pack(anchor="w", pady=(6, 0))

    def on_show(self):
        pass

    # ── Import ────────────────────────────────────────────────────────────────

    def _browse_import(self):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        self._import_path_var.set(path)
        self._import_btn.config(state="normal")
        self._preview_import(path)

    def _preview_import(self, path: str):
        try:
            import openpyxl
            cfg = _load_cfg()["import"]
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if cfg["sheet"] not in wb.sheetnames:
                self._import_status.config(
                    text=f"Sheet \"{cfg['sheet']}\" not found. Available: {', '.join(wb.sheetnames)}",
                    fg="#C00000")
                self._import_btn.config(state="disabled")
                return

            ws = wb[cfg["sheet"]]
            col_map = self._build_col_index(ws, cfg)
            total, no_price = 0, 0
            for row in ws.iter_rows(min_row=cfg["header_row"] + 1, values_only=True):
                title_idx = col_map.get("title")
                if title_idx is None or not row[title_idx]:
                    continue
                total += 1
                price_idx = col_map.get("price")
                if price_idx is None or not row[price_idx]:
                    no_price += 1
            wb.close()

            warn = f"  |  {no_price} rows missing price (will default to {cfg['default_price']})" if no_price else ""
            self._import_status.config(
                text=f"Ready: {total} products found{warn}", fg="#007000")
        except Exception as exc:
            self._import_status.config(text=f"Preview error: {exc}", fg="#C00000")

    def _run_import(self):
        path = self._import_path_var.get()
        if not path:
            return
        replace = self._replace_var.get()
        if replace and not messagebox.askyesno(
            "Replace products",
            "This will delete ALL existing products before importing.\n\nContinue?",
            parent=self,
        ):
            return
        self._import_btn.config(state="disabled")
        self._import_status.config(text="Importing…", fg=FG_MUTED)
        threading.Thread(target=self._do_import, args=(path, replace), daemon=True).start()

    def _do_import(self, path: str, replace: bool):
        try:
            import openpyxl
            cfg = _load_cfg()["import"]
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[cfg["sheet"]]
            col_map = self._build_col_index(ws, cfg)
            default_price = float(cfg.get("default_price", 0.0))

            rows = []
            for row in ws.iter_rows(min_row=cfg["header_row"] + 1, values_only=True):
                t_idx = col_map.get("title")
                if t_idx is None or not row[t_idx]:
                    continue

                def _str(key, r=row):
                    idx = col_map.get(key)
                    return str(r[idx]).strip() if idx is not None and r[idx] is not None else ""

                def _int(key, r=row):
                    idx = col_map.get(key)
                    try:
                        return int(r[idx]) if idx is not None and r[idx] is not None else None
                    except (TypeError, ValueError):
                        return None

                def _float(key, r=row):
                    idx = col_map.get(key)
                    try:
                        return float(r[idx]) if idx is not None and r[idx] is not None else None
                    except (TypeError, ValueError):
                        return None

                price = _float("price")
                barcode_val = _str("barcode") or None
                rows.append({
                    "id":        _int("id"),
                    "title":     _str("title"),
                    "author":    _str("author"),
                    "publisher": _str("publisher"),
                    "webstore":  _str("webstore"),
                    "location":  _str("location"),
                    "price":     price if price is not None else default_price,
                    "stock":     _int("stock") or 0,
                    "storage":   _int("storage"),
                    "barcode":   barcode_val,
                })
            wb.close()

            from db.database import get_connection
            conn = get_connection()

            if replace:
                conn.execute("DELETE FROM products")
                conn.commit()
                self._log_msg("Cleared existing products.")

            imported, skipped = 0, 0
            for r in rows:
                try:
                    product_model.import_product(
                        r["title"], r["stock"], r["price"],
                        product_id=r["id"],
                        barcode=r["barcode"],
                        author=r["author"], publisher=r["publisher"],
                        webstore=r["webstore"], location=r["location"],
                        storage=r["storage"],
                    )
                    imported += 1
                except Exception as e:
                    self._log_msg(f"  Skipped \"{r['title'][:40]}\": {e}")
                    skipped += 1

            # Reset autoincrement so new products get IDs above the max imported ID
            conn.execute(
                "UPDATE sqlite_sequence SET seq = (SELECT MAX(id) FROM products)"
                " WHERE name = 'products'"
            )
            conn.commit()

            self.after(0, lambda: self._import_done(imported, skipped))
        except Exception as exc:
            self.after(0, lambda: self._import_error(str(exc)))

    def _import_done(self, imported: int, skipped: int):
        msg = f"Done: {imported} imported, {skipped} skipped."
        self._import_status.config(text=msg, fg="#007000")
        self._import_btn.config(state="normal")
        self._log_msg(msg)

    def _import_error(self, msg: str):
        self._import_status.config(text=f"Error: {msg}", fg="#C00000")
        self._import_btn.config(state="normal")
        self._log_msg(f"Import error: {msg}")

    # ── Export ────────────────────────────────────────────────────────────────

    def _run_export(self):
        path = filedialog.asksaveasfilename(
            title="Save Excel file",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="products_export.xlsx",
        )
        if not path:
            return
        try:
            self._do_export(path)
        except Exception as exc:
            self._export_status.config(text=f"Error: {exc}", fg="#C00000")
            self._log_msg(f"Export error: {exc}")

    def _do_export(self, path: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        cfg = _load_cfg()["export"]
        col_cfg = cfg["columns"]   # field → header label
        fields  = list(col_cfg.keys())
        headers = list(col_cfg.values())

        products = product_model.get_all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = cfg.get("sheet", "Products")

        hdr_fill = PatternFill("solid", fgColor="2E6DA4")
        hdr_font = Font(color="FFFFFF", bold=True)
        for col_idx, label in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        for r_idx, p in enumerate(products, 2):
            for c_idx, field in enumerate(fields, 1):
                ws.cell(row=r_idx, column=c_idx, value=p.get(field))

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        wb.save(path)
        msg = f"Exported {len(products)} products to {os.path.basename(path)}"
        self._export_status.config(text=msg, fg="#007000")
        self._log_msg(msg)

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _build_col_index(self, ws, cfg: dict) -> dict[str, int]:
        """Map field name → 0-based column index by matching header row values."""
        header_row = list(ws.iter_rows(
            min_row=cfg["header_row"], max_row=cfg["header_row"],
            values_only=True,
        ))[0]
        name_to_idx = {
            str(v).strip().upper(): i
            for i, v in enumerate(header_row)
            if v is not None
        }
        result = {}
        for field, col_label in cfg["columns"].items():
            key = col_label.strip().upper()
            if key in name_to_idx:
                result[field] = name_to_idx[key]
        return result

    def _log_msg(self, text: str):
        def _do():
            self._log.config(state="normal")
            self._log.insert("end", text + "\n")
            self._log.see("end")
            self._log.config(state="disabled")
        self.after(0, _do)
